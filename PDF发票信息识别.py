import pdfplumber
import re
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from tkinter import Tk
from tkinter.filedialog import askdirectory
import subprocess

# 使用tkinter选择文件夹
root = Tk()
root.withdraw()  # 隐藏主窗口
pdf_folder = askdirectory(title="选择PDF文件所在的文件夹")  # 弹出文件夹选择对话框

# 正则表达式模式
invoice_date_pattern1 = r'(\d\s*\d\s*\d\s*\d\s*年\s*\d\s*\d?\s*月\s*\d\s*\d?\s*日)'  # 匹配开票日期，格式为****年**月**日，年、月、日字符前面可能存在空格
invoice_date_pattern2 = r'(\d{4}\s+\d{1,2}\s+\d{1,2})'  # 匹配开票日期，格式为**** ** **，年、月、日之间可能存在空格
tax_rate_pattern = r'(\d+(\.\d+)?%)'  # 匹配税率
amount_pattern = r'(\d+\.\d{2})'  # 匹配金额，两位小数点的数字
invoice_subject_pattern = r'^\*([^\dA-Za-z]+)\*([^\s]+)'  # 匹配*字符后面和下面的所有字符，直到空白字符结束，且两个*之间不会有数字和字母
name_pattern_colon = r'称\s*[:：]{1,3}\s*([\u4e00-\u9fa5]+)'  # 匹配“称：”后的汉字内容，冒号可以是半角或全角，且可能有1到3个冒号，称和冒号之间可能有空格
name_pattern_no_colon = r'称\s*([\u4e00-\u9fa5]+)'  # 匹配“称”后的汉字内容
company_pattern = r'([\u4e00-\u9fa5]+公司)\s+([^\s]+)'  # 匹配“公司”+一个空格的位置
invoice_number_pattern = r'发票\s*[\u4e00-\u9fa5]*\s*号码'  # 匹配“发票+号码”的字符位置，发票和号码中间可能有其他字符

# 初始化一个空的列表，用于存储所有提取的信息
all_invoice_data = []

# 遍历文件夹中的所有PDF文件，包括子文件夹
for root, dirs, files in os.walk(pdf_folder):
    for filename in files:
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(root, filename)
            
            # 打开并提取PDF中的文本
            with pdfplumber.open(pdf_path) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text

                # 使用正则表达式优先提取“称：”后的汉字内容
                names_colon = re.findall(name_pattern_colon, text)  # 匹配所有“称：”后面的汉字内容
                names_no_colon = re.findall(name_pattern_no_colon, text)  # 匹配所有“称”后面的汉字内容

                # 优先使用“称：”识别的结果
                if len(names_colon) >= 2 and len(names_colon[0].strip()) >= 3 and len(names_colon[1].strip()) >= 3:
                    name1 = names_colon[0].strip()
                    name2 = names_colon[1].strip()
                elif len(names_colon) == 1 and len(names_colon[0].strip()) >= 3:
                    name1 = names_colon[0].strip()
                    name2 = None
                else:
                    # 如果没有找到“称：”或识别的内容少于3个字符，则使用“称”识别的结果
                    if len(names_no_colon) >= 2 and len(names_no_colon[0].strip()) >= 3 and len(names_no_colon[1].strip()) >= 3:
                        name1 = names_no_colon[0].strip()
                        name2 = names_no_colon[1].strip()
                    elif len(names_no_colon) == 1 and len(names_no_colon[0].strip()) >= 3:
                        name1 = names_no_colon[0].strip()
                        name2 = None
                    else:
                        name1 = None
                        name2 = None

                # 如果前两种方法未识别到名称或识别的内容少于3个字符，使用识别“公司”+一个空格的位置的方法
                if not name1 or not name2 or len(name1) < 3 or len(name2) < 3:
                    company_match = re.search(company_pattern, text)
                    if company_match:
                        name1 = company_match.group(1).strip()
                        name2 = company_match.group(2).strip()

                # 使用正则表达式提取开票日期
                invoice_date_match = re.search(invoice_date_pattern1, text)
                if not invoice_date_match:
                    invoice_date_match = re.search(invoice_date_pattern2, text)
                    if invoice_date_match:
                        invoice_date = invoice_date_match.group(1).replace(" ", "-")
                else:
                    invoice_date = invoice_date_match.group(1).replace(" ", "")

                # 提取发票号码，通过识别“发票+号码”的字符位置，发票和号码中间可能有其他字符
                invoice_number = None
                invoice_number_match = re.search(invoice_number_pattern, text)
                if invoice_number_match:
                    lines = text.split('\n')
                    for line in lines:
                        if invoice_number_match.group() in line:
                            number_match = re.search(r'\d+', line)
                            if number_match:
                                invoice_number = number_match.group()
                                break

                # 如果第一种方法未识别到发票号码，使用第二种方法，通过识别开票日期上方的纯数字字符串
                if not invoice_number and invoice_date:
                    date_position = text.find(invoice_date)
                    if date_position != -1:
                        lines = text[:date_position].split('\n')
                        for line in reversed(lines):
                            if re.match(r'^\d+$', line.strip()):
                                invoice_number = line.strip()
                                break

                # 提取税率
                tax_rate = re.search(tax_rate_pattern, text)
                tax_rate = tax_rate.group(1) if tax_rate else None

                # 提取发票金额，通过识别“价税合计”字符，并提取该行中的两位小数点的数字
                invoice_amount = None
                for line in text.split('\n'):
                    if "价税合计" in line:
                        amount_match = re.search(amount_pattern, line)
                        if amount_match:
                            invoice_amount = float(amount_match.group(1))
                            break

                # 提取不含税金额和税额，通过识别“合计”或“合 计”字符，并提取该行中的两位小数点的数字
                amount_excl_tax = None
                tax_amount = None
                for line in text.split('\n'):
                    if re.search(r'合\s*计', line):
                        amounts = re.findall(amount_pattern, line)
                        if len(amounts) >= 2:
                            amount_excl_tax = float(amounts[0])  # 靠左边的为不含税金额
                            tax_amount = float(amounts[1])  # 靠右边的为税额
                            break

                # 如果上述方法未识别到不含税金额和税额，使用备用方法
                if amount_excl_tax is None or tax_amount is None:
                    amounts = re.findall(amount_pattern, text)
                    if len(amounts) >= 2:
                        amounts = sorted([float(amount) for amount in amounts], reverse=True)
                        if len(amounts) > 1:
                            amount_excl_tax = amounts[1]  # 数值第二大的为不含税金额
                            tax_amount = invoice_amount - amount_excl_tax if invoice_amount else None

                # 提取发票科目，按照“*”+字符+“*”+字符的规律来识别，识别到空格结束，第一个*为所在行的第一个字符
                invoice_subject_match = re.findall(invoice_subject_pattern, text, re.MULTILINE)
                invoice_subject = invoice_subject_match[0][1].strip() if invoice_subject_match else None

                # 提取发票类型，通过识别PDF中是否包含“普通发票”
                if "普通发票" in text:
                    invoice_type = "增值税普通发票"
                else:
                    invoice_type = "增值税专用发票"

                # 将信息添加到数据列表
                all_invoice_data.append({
                    "文件名": pdf_path,  # 添加文件路径
                    "发票号码": invoice_number,
                    "开票日期": invoice_date,
                    "购买方": name1,
                    "销售方": name2,
                    "税率": tax_rate,
                    "发票金额": invoice_amount,
                    "不含税金额": amount_excl_tax,
                    "税额": tax_amount,
                    "发票科目": invoice_subject,
                    "发票类型": invoice_type
                })

# 将所有数据转换为 DataFrame
df = pd.DataFrame(all_invoice_data)

# 导出数据到 Excel 文件
output_file = os.path.join(pdf_folder, "发票汇总.xlsx")
df.to_excel(output_file, index=False, engine='openpyxl')

# 调整Excel列宽并添加超链接
wb = load_workbook(output_file)
ws = wb.active

for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # 获取列字母
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# 添加超链接
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    for cell in row:
        file_path = cell.value
        cell.hyperlink = Hyperlink(ref=cell.coordinate, target=file_path)
        cell.font = Font(color="0000FF", underline="single")

wb.save(output_file)

# 自动打开生成的Excel文件
subprocess.Popen(['start', output_file], shell=True)

print(f"所有提取的发票信息已导出到: {output_file}")